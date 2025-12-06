#!/usr/bin/env python3
"""
BOB Data Exporters v4.3.1

Export extracted business data to multiple formats:
- JSON (default)
- CSV (spreadsheet compatible)
- SQLite (database)
- Excel (requires openpyxl)

Usage:
    from bob.utils.exporters import export_to_csv, export_to_excel, export_to_sqlite
    
    export_to_csv(businesses, 'output.csv')
    export_to_excel(businesses, 'output.xlsx')
    export_to_sqlite(businesses, 'output.db')
"""

import csv
import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, TYPE_CHECKING

# Type hints for optional openpyxl (suppresses Pylance warning)
if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================================
# JSON EXPORT
# ============================================================================

def export_to_json(
    data: List[Dict[str, Any]],
    filepath: str,
    include_metadata: bool = True,
    indent: int = 2
) -> str:
    """
    Export business data to JSON file.
    
    Args:
        data: List of business dictionaries
        filepath: Output file path
        include_metadata: Add extraction metadata
        indent: JSON indentation (default: 2)
    
    Returns:
        Path to created file
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    if include_metadata:
        output = {
            "metadata": {
                "exported_at": datetime.now().isoformat(),
                "exporter_version": "4.3.1",
                "format": "json",
                "total_records": len(data),
            },
            "businesses": data
        }
    else:
        output = data
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=indent, ensure_ascii=False, default=str)
    
    print(f"✅ Exported {len(data)} businesses to {filepath}")
    return str(filepath)


# ============================================================================
# CSV EXPORT
# ============================================================================

# Fields to export in CSV (flat structure, no nested data)
CSV_FIELDS = [
    'name',
    'phone',
    'address',
    'website',
    'rating',
    'reviews_count',
    'category',
    'latitude',
    'longitude',
    'place_id_hex',
    'cid',
    'quality_score',
    'extraction_method',
]


def export_to_csv(
    data: List[Dict[str, Any]],
    filepath: str,
    fields: Optional[List[str]] = None,
    include_header: bool = True
) -> str:
    """
    Export business data to CSV file.
    
    Args:
        data: List of business dictionaries
        filepath: Output file path
        fields: List of fields to export (default: CSV_FIELDS)
        include_header: Include header row
    
    Returns:
        Path to created file
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    fields = fields or CSV_FIELDS
    
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        
        if include_header:
            writer.writeheader()
        
        for business in data:
            # Flatten nested data for CSV
            row = {}
            for field in fields:
                value = business.get(field, '')
                # Convert lists/dicts to string
                if isinstance(value, (list, dict)):
                    value = json.dumps(value)
                row[field] = value
            writer.writerow(row)
    
    print(f"✅ Exported {len(data)} businesses to {filepath}")
    return str(filepath)


# ============================================================================
# SQLITE EXPORT
# ============================================================================

def export_to_sqlite(
    data: List[Dict[str, Any]],
    filepath: str,
    table_name: str = 'businesses'
) -> str:
    """
    Export business data to SQLite database.
    
    Args:
        data: List of business dictionaries
        filepath: Output database file path
        table_name: Name of the table (default: 'businesses')
    
    Returns:
        Path to created file
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    conn = sqlite3.connect(filepath)
    cursor = conn.cursor()
    
    # Create table
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            phone TEXT,
            address TEXT,
            website TEXT,
            rating REAL,
            reviews_count INTEGER,
            category TEXT,
            latitude REAL,
            longitude REAL,
            place_id_hex TEXT UNIQUE,
            cid TEXT,
            quality_score INTEGER,
            images TEXT,
            reviews TEXT,
            full_data TEXT,
            exported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    # Insert data
    for business in data:
        try:
            cursor.execute(f"""
                INSERT OR REPLACE INTO {table_name} (
                    name, phone, address, website, rating, reviews_count,
                    category, latitude, longitude, place_id_hex, cid,
                    quality_score, images, reviews, full_data
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                business.get('name'),
                business.get('phone'),
                business.get('address'),
                business.get('website'),
                business.get('rating'),
                business.get('reviews_count'),
                business.get('category'),
                business.get('latitude'),
                business.get('longitude'),
                business.get('place_id_hex'),
                business.get('cid'),
                business.get('quality_score'),
                json.dumps(business.get('images', [])),
                json.dumps(business.get('reviews', [])),
                json.dumps(business)
            ))
        except sqlite3.IntegrityError:
            # Skip duplicates
            pass
    
    conn.commit()
    conn.close()
    
    print(f"✅ Exported {len(data)} businesses to {filepath}")
    return str(filepath)


# ============================================================================
# EXCEL EXPORT (Optional - requires openpyxl)
# ============================================================================

def export_to_excel(
    data: List[Dict[str, Any]],
    filepath: str,
    sheet_name: str = 'Businesses',
    fields: Optional[List[str]] = None
) -> str:
    """
    Export business data to Excel file.
    
    Requires: pip install openpyxl
    
    Args:
        data: List of business dictionaries
        filepath: Output file path (.xlsx)
        sheet_name: Name of the worksheet
        fields: List of fields to export (default: CSV_FIELDS)
    
    Returns:
        Path to created file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        )
    
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    fields = fields or CSV_FIELDS
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Write header
    for col, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col, value=field.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_num, business in enumerate(data, 2):
        for col, field in enumerate(fields, 1):
            value = business.get(field, '')
            if isinstance(value, (list, dict)):
                value = json.dumps(value)
            ws.cell(row=row_num, column=col, value=value)
    
    # Auto-width columns
    for col in range(1, len(fields) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Name column wider
    ws.column_dimensions['A'].width = 30
    # Address column wider
    if 'address' in fields:
        addr_col = fields.index('address') + 1
        ws.column_dimensions[chr(64 + addr_col)].width = 40
    
    wb.save(filepath)
    print(f"✅ Exported {len(data)} businesses to {filepath}")
    return str(filepath)


# ============================================================================
# MULTI-FORMAT EXPORT
# ============================================================================

def export_all_formats(
    data: List[Dict[str, Any]],
    base_path: str,
    formats: Optional[List[str]] = None
) -> Dict[str, str]:
    """
    Export to multiple formats at once.
    
    Args:
        data: List of business dictionaries
        base_path: Base path without extension (e.g., 'output/businesses')
        formats: List of formats ['json', 'csv', 'sqlite', 'excel']
    
    Returns:
        Dictionary mapping format to file path
    """
    formats = formats or ['json', 'csv', 'sqlite']
    results = {}
    
    base = Path(base_path)
    
    if 'json' in formats:
        results['json'] = export_to_json(data, f"{base}.json")
    
    if 'csv' in formats:
        results['csv'] = export_to_csv(data, f"{base}.csv")
    
    if 'sqlite' in formats:
        results['sqlite'] = export_to_sqlite(data, f"{base}.db")
    
    if 'excel' in formats:
        try:
            results['excel'] = export_to_excel(data, f"{base}.xlsx")
        except ImportError as e:
            print(f"⚠️ Excel export skipped: {e}")
    
    return results


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def load_json_data(filepath: str) -> List[Dict[str, Any]]:
    """Load business data from JSON file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Handle both formats: list or dict with 'businesses' key
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        return data.get('businesses', [])
    return []


def convert_json_to_csv(json_path: str, csv_path: Optional[str] = None) -> str:
    """Convert existing JSON file to CSV."""
    data = load_json_data(json_path)
    csv_path = csv_path or json_path.replace('.json', '.csv')
    return export_to_csv(data, csv_path)


def convert_json_to_sqlite(json_path: str, db_path: Optional[str] = None) -> str:
    """Convert existing JSON file to SQLite."""
    data = load_json_data(json_path)
    db_path = db_path or json_path.replace('.json', '.db')
    return export_to_sqlite(data, db_path)


# ============================================================================
# CLI INTERFACE
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="BOB Data Exporter")
    parser.add_argument('input', help='Input JSON file')
    parser.add_argument('--format', '-f', choices=['csv', 'sqlite', 'excel', 'all'],
                        default='csv', help='Output format')
    parser.add_argument('--output', '-o', help='Output file path')
    
    args = parser.parse_args()
    
    data = load_json_data(args.input)
    
    if args.format == 'all':
        base = args.output or args.input.replace('.json', '')
        export_all_formats(data, base)
    elif args.format == 'csv':
        output = args.output or args.input.replace('.json', '.csv')
        export_to_csv(data, output)
    elif args.format == 'sqlite':
        output = args.output or args.input.replace('.json', '.db')
        export_to_sqlite(data, output)
    elif args.format == 'excel':
        output = args.output or args.input.replace('.json', '.xlsx')
        export_to_excel(data, output)
